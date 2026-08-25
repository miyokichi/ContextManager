"""Excel extractor: sheet names, used range, a bounded peek at values, and
merged-cell info. Never loads a whole workbook into the LLM prompt - just
enough for the AI Analyzer to guess what's where in each sheet."""
from __future__ import annotations

from ..models import ExtractedDocument, ExtractedSection
from .base import ExtractionError


def used_range(ws) -> str:
    """Best-effort used-range string. `ws.dimensions` isn't available on
    read-only worksheets in every openpyxl version, so fall back to
    max_row/max_column (which are)."""
    try:
        return ws.dimensions
    except AttributeError:
        pass
    try:
        from openpyxl.utils import get_column_letter

        if ws.max_row and ws.max_column:
            return f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass
    return "unknown"


def extract(path, max_rows: int = 15, max_cols: int = 10, max_chars_per_sheet: int = 3000) -> ExtractedDocument:
    try:
        import openpyxl
    except ImportError as e:
        raise ExtractionError("openpyxl is required to read Excel files") from e

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ExtractionError(f"failed to open workbook: {e}") from e

    sections = []
    try:
        for ws in wb.worksheets:
            lines = [f"used_range={used_range(ws)}"]
            try:
                merged = list(ws.merged_cells.ranges)[:10]
            except Exception:
                merged = []
            if merged:
                lines.append("merged_cells: " + ", ".join(str(m) for m in merged))
            for row in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
                if any(c is not None for c in row):
                    lines.append(" | ".join("" if c is None else str(c) for c in row))
            text = "\n".join(lines)
            if len(text) > max_chars_per_sheet:
                text = text[:max_chars_per_sheet] + "\n...(truncated)"
            sections.append(ExtractedSection(location=ws.title, text=text))
    finally:
        wb.close()

    return ExtractedDocument(kind="xlsx", sections=sections)
